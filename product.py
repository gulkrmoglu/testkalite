from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
import time
import pytest
import openpyxl

options = Options()
options.add_argument("start-maximized")
class Test_Sauce:
    def __init__(self):
        self.driver=webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        self.driver.maximize_window()
        self.driver.get("https://www.saucedemo.com/")
       
    def test_valid_login(self):
        userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"user-name")))
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"password")))
        userNameInput.send_keys("standard_user")
        passwordInput.send_keys("secret_sauce")
        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"login-button")))
        loginButton.click()
    
    def test_products_filter(self):
        
       filter_dropdown = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CLASS_NAME, "product_sort_container")))
       time.sleep(3)
       select = Select(filter_dropdown)
       time.sleep(3)
       select.select_by_visible_text("Price (low to high)")
       time.sleep(3)

      
    def test_product_details(self): 
        
        product_link = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CSS_SELECTOR, ".inventory_item:nth-child(1) .inventory_item_name")))
        product_link.click()
        product_name = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CLASS_NAME, "inventory_details_name")))
        assert product_name != ""
        product_price = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CLASS_NAME, "inventory_details_price")))
        assert product_price != ""
        product_description = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CLASS_NAME, "inventory_details_desc")))
        assert product_description != ""
        time.sleep(5)

    def getData():
        excelFile=openpyxl.load_workbook("data/test_add_to_cart.xlsx")
        selectedSheet=excelFile["Sayfa1"]

        totalRows=selectedSheet.max_row
        data=[]
        for i in range(1, totalRows+1):
               product_name=selectedSheet.cell(i,1).value
               product_price=selectedSheet.cell(i,2).value
               tupleData=(product_name, product_price)
               data.append(tupleData)

        return data   
       
    @pytest.mark.parametrize("product_name, product_price",getData()) 
    def test_add_to_cart(self, cart_product_name, cart_product_price):
        add_to_cart_button = WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.ID, "add-to-cart")))
        add_to_cart_button.click()
        cart_icon = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CLASS_NAME, "shopping_cart_link")))
        cart_icon.click()
        cart_product_name = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CLASS_NAME, "inventory_item_name")))
        assert cart_product_name != ""
        cart_product_price = WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CLASS_NAME, "inventory_item_price")))
        assert cart_product_price != ""

        time.sleep(3)
        
testclass=Test_Sauce()
testclass.test_valid_login() 
testclass.test_products_filter()
testclass.test_product_details()
product_name = "Product Name"
product_price = "Product Price"
testclass.test_add_to_cart(product_name, product_price)




    

